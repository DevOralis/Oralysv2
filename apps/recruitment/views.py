from django.shortcuts import render, redirect
from django.contrib import messages
from django.utils import timezone
from django.template.loader import render_to_string
from datetime import datetime
from .models import Offer, Candidate, Application, Interview, Pointage, Evaluation
from apps.hr.models import Department, Employee
from django.db.models import Q
from django.http import JsonResponse, HttpResponse, FileResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_GET
import json
import openpyxl
import csv
import io
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from django.db.models import Count, Avg, F, DurationField, ExpressionWrapper
from datetime import timedelta
from calendar import monthrange
from django.db.models.functions import TruncMonth
from django.core.paginator import Paginator
from django.db.models.functions import ExtractHour, ExtractMinute, Extract
from PIL import Image

from datetime import datetime


def recruitment_home(request):
    if request.method == 'POST':
        action = request.POST.get('action', 'add')
        
        if action == 'add':
            # Logique d'ajout d'offre
            try:
                # Récupération des données du formulaire
                title = request.POST.get('title')
                description = request.POST.get('description', '')
                publication_date = request.POST.get('publication_date')
                end_date = request.POST.get('end_date')
                skills = request.POST.get('skills', '')
                profile = request.POST.get('profile', '')
                department_id = request.POST.get('department')
                try:
                    positions_available = int(request.POST.get('positions_available', 1))
                except (ValueError, TypeError):
                    positions_available = 1
                # Validation des champs requis
                if not title or not publication_date or not end_date:
                    messages.error(request, 'Veuillez remplir tous les champs obligatoires.')
                    return redirect('recruitment_home')
                
                # Conversion des dates
                try:
                    pub_date = datetime.strptime(publication_date, '%Y-%m-%d').date()
                    end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
                except ValueError:
                    messages.error(request, 'Format de date invalide.')
                    return redirect('recruitment_home')
                
                # Récupération du département
                department = None
                if department_id:
                    try:
                        department = Department.objects.get(id=department_id)
                    except Department.DoesNotExist:
                        department = None
                
                # Création de l'offre
                offer = Offer.objects.create(
                    title=title,
                    description=description,
                    publication_date=pub_date,
                    end_date=end_date_obj,
                    skills=skills,
                    profile=profile,
                    department=department,
                    positions_available=positions_available
                )
                
                messages.success(request, f'Offre "{title}" ajoutée avec succès!')
                return redirect('recruitment_home')
                
            except Exception as e:
                messages.error(request, f'Erreur lors de l\'ajout de l\'offre: {str(e)}')
                return redirect('recruitment_home')
        
        elif action == 'edit':
            # Logique de modification d'offre
            try:
                offer_id = request.POST.get('offer_id')
                title = request.POST.get('title')
                description = request.POST.get('description', '')
                publication_date = request.POST.get('publication_date')
                end_date = request.POST.get('end_date')
                skills = request.POST.get('skills', '')
                profile = request.POST.get('profile', '')
                department_id = request.POST.get('department')
                try:
                    positions_available = int(request.POST.get('positions_available', 1))
                except (ValueError, TypeError):
                    positions_available = 1
                # Validation des champs requis
                if not offer_id or not title or not publication_date or not end_date:
                    messages.error(request, 'Veuillez remplir tous les champs obligatoires.')
                    return redirect('recruitment_home')
                
                # Récupération de l'offre
                try:
                    offer = Offer.objects.get(id=offer_id)
                except Offer.DoesNotExist:
                    messages.error(request, 'Offre introuvable.')
                    return redirect('recruitment_home')
                
                # Conversion des dates
                try:
                    pub_date = datetime.strptime(publication_date, '%Y-%m-%d').date()
                    end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
                except ValueError:
                    messages.error(request, 'Format de date invalide.')
                    return redirect('recruitment_home')
                
                # Récupération du département
                department = None
                if department_id:
                    try:
                        department = Department.objects.get(id=department_id)
                    except Department.DoesNotExist:
                        department = None
                
                # Mise à jour de l'offre
                offer.title = title
                offer.description = description
                offer.publication_date = pub_date
                offer.end_date = end_date_obj
                offer.skills = skills
                offer.profile = profile
                offer.department = department
                offer.positions_available = positions_available
                offer.save()
                
                messages.success(request, f'Offre "{title}" modifiée avec succès!')
                return redirect('recruitment_home')
                
            except Exception as e:
                messages.error(request, f'Erreur lors de la modification de l\'offre: {str(e)}')
                return redirect('recruitment_home')
        
        elif action == 'delete':
            # Logique de suppression d'offre
            try:
                offer_id = request.POST.get('offer_id')
                
                if not offer_id:
                    messages.error(request, 'ID de l\'offre manquant.')
                    return redirect('recruitment_home')
                
                # Récupération et suppression de l'offre
                try:
                    offer = Offer.objects.get(id=offer_id)
                    offer_title = offer.title
                    offer.delete()
                    messages.success(request, f'Offre "{offer_title}" supprimée avec succès!')
                except Offer.DoesNotExist:
                    messages.error(request, 'Offre introuvable.')
                
                return redirect('recruitment_home')
                
            except Exception as e:
                messages.error(request, f'Erreur lors de la suppression de l\'offre: {str(e)}')
                return redirect('recruitment_home')
        
        elif action == 'unarchive':
            # Réactivation d'une offre archivée
            try:
                offer_id = request.POST.get('offer_id')
                if not offer_id:
                    messages.error(request, "ID de l'offre manquant.")
                    return redirect('recruitment_home')
                offer = Offer.objects.get(id=offer_id)
                offer.archived = False
                offer.save()
                messages.success(request, f'Offre "{offer.title}" réactivée avec succès!')
                return redirect('recruitment_home')
            except Offer.DoesNotExist:
                messages.error(request, "Offre introuvable.")
                return redirect('recruitment_home')
    
    # GET request - affichage de la liste
    try:
        query = request.GET.get('q', '').strip()
        department_id = request.GET.get('department', '').strip()
        print(f"Recherche reçue: '{query}'")
        
        offers = Offer.objects.filter(archived=False).annotate(num_applications=Count('application')).order_by('title')
        archived_offers = Offer.objects.filter(archived=True).annotate(num_applications=Count('application')).order_by('title')
        
        # Recherche pour les offres actives
        if query:
            offers = offers.filter(title__icontains=query)
            print(f"Offres trouvées: {offers.count()}")
        if department_id:
            offers = offers.filter(department_id=department_id)
        offers = offers.order_by('-publication_date')
        
        # Recherche pour les offres archivées
        archived_query = request.GET.get('archived_q', '').strip()
        archived_department_id = request.GET.get('archived_department', '').strip()
        
        if archived_query:
            archived_offers = archived_offers.filter(title__icontains=archived_query)
        if archived_department_id:
            archived_offers = archived_offers.filter(department_id=archived_department_id)
        archived_offers = archived_offers.order_by('-publication_date')
        
        # Ajout pagination
        paginator = Paginator(offers, 5)  # 5 offres par page
        page_number = request.GET.get('page')
        page_obj = paginator.get_page(page_number)
        departments = Department.objects.all().order_by('name')
        context = {
            'offers': offers,
            'page_obj': page_obj,
            'query': query,
            'today': timezone.now().date(),
            'departments': departments,
            'selected_department': department_id,
            'archived_offers': archived_offers,
            'archived_query': archived_query,
            'archived_selected_department': archived_department_id,
        }
        
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            print("Rendu HTML direct pour AJAX")
            from django.http import HttpResponse
            html_content = render_to_string('offer.html', context)
            
            # Vérifier si c'est une recherche pour les offres archivées (basé sur la présence des paramètres)
            if 'archived_q' in request.GET or 'archived_department' in request.GET:
                # Extraire seulement la partie du tableau des offres archivées
                import re
                table_match = re.search(r'<tbody id="archived-offers-table">(.*?)</tbody>', html_content, re.DOTALL)
                if table_match:
                    table_content = table_match.group(1)
                    return HttpResponse(table_content)
                else:
                    return HttpResponse('<tr><td colspan="8" class="text-center text-muted">Aucune offre archivée trouvée.</td></tr>')
            else:
                # Extraire seulement la partie du tableau des offres actives
                import re
                table_match = re.search(r'<tbody id="offers-table">(.*?)</tbody>', html_content, re.DOTALL)
                if table_match:
                    table_content = table_match.group(1)
                    return HttpResponse(table_content)
                else:
                    return HttpResponse('<tr><td colspan="5" class="text-center text-danger">Erreur: Impossible de trouver le tableau</td></tr>')
        
        print("Rendu template complet")
        return render(request, 'offer.html', context)
    except Exception as e:
        print(f"Erreur dans la vue: {str(e)}")
        import traceback
        traceback.print_exc()
        if request.headers.get('x-requested-with') == 'XMLHttpRequest':
            from django.http import JsonResponse
            return JsonResponse({'error': str(e)}, status=500)
        raise


def recruitment_candidates(request):
    if request.method == 'POST':
        action = request.POST.get('action', 'add')
        try:
            if action == 'add':
                # Extraction automatique depuis le CV si fourni
                cv_file = request.FILES.get('cv')
                extracted = {}
                if cv_file:
                    # Extraction du texte du CV (PDF ou image)
                    import PyPDF2
                    text = ''
                    if cv_file.name.lower().endswith('.pdf'):
                        try:
                            reader = PyPDF2.PdfReader(cv_file)
                            for page in reader.pages:
                                text += page.extract_text() or ''
                        except Exception:
                            text = ''
                    # Ajoute ici l'extraction image si besoin
                    if text:
                        extracted = extract_cv_data(text)
                last_name = request.POST.get('last_name') or extracted.get('last_name', '')
                first_name = request.POST.get('first_name') or extracted.get('first_name', '')
                email = request.POST.get('email')
                phone = request.POST.get('phone') or extracted.get('phone', '')
                # Normalize Moroccan phone number to +212XXXXXXXXX when provided
                def _normalize_moroccan_phone(raw):
                    try:
                        if not raw:
                            return ''
                        digits = re.sub(r'\D', '', str(raw))
                        if not digits:
                            return ''
                        if digits.startswith('00212'):
                            digits = digits[2:]
                        if digits.startswith('212') and len(digits) >= 12:
                            return '+212' + digits[3:12]
                        if digits.startswith('0') and len(digits) >= 10:
                            return '+212' + digits[1:10]
                        return ''
                    except Exception:
                        return ''
                phone = _normalize_moroccan_phone(phone)
                birth_date = request.POST.get('birth_date')
                address = request.POST.get('address') or extracted.get('address', '')
                gender = request.POST.get('gender')
                photo = request.FILES.get('photo')
                linkedin_profile = request.POST.get('linkedin_profile')
                if not last_name or not first_name or not email:
                    messages.error(request, 'Veuillez remplir tous les champs obligatoires.')
                    return redirect('recruitment_candidates')
                candidate = Candidate(
                    last_name=last_name,
                    first_name=first_name,
                    email=email,
                    phone=phone,
                    address=address,
                    gender=gender,
                    linkedin_profile=linkedin_profile
                )
                if birth_date:
                    candidate.birth_date = birth_date
                if photo:
                    candidate.photo = photo
                candidate.save()
                messages.success(request, f'Candidat {first_name} {last_name} ajouté avec succès!')
                return redirect('recruitment_candidates')
            elif action == 'edit':
                candidate_id = request.POST.get('candidate_id')
                last_name = request.POST.get('last_name')
                first_name = request.POST.get('first_name')
                email = request.POST.get('email')
                phone = request.POST.get('phone')
                birth_date = request.POST.get('birth_date')
                address = request.POST.get('address')
                gender = request.POST.get('gender')
                photo = request.FILES.get('photo')
                linkedin_profile = request.POST.get('linkedin_profile')
                if not candidate_id or not last_name or not first_name or not email:
                    messages.error(request, 'Veuillez remplir tous les champs obligatoires.')
                    return redirect('recruitment_candidates')
                try:
                    candidate = Candidate.objects.get(id=candidate_id)
                except Candidate.DoesNotExist:
                    messages.error(request, 'Candidat introuvable.')
                    return redirect('recruitment_candidates')
                candidate.last_name = last_name
                candidate.first_name = first_name
                candidate.email = email
                # Normalize on edit as well
                def _normalize_moroccan_phone_edit(raw):
                    try:
                        if not raw:
                            return ''
                        digits = re.sub(r'\D', '', str(raw))
                        if not digits:
                            return ''
                        if digits.startswith('00212'):
                            digits = digits[2:]
                        if digits.startswith('212') and len(digits) >= 12:
                            return '+212' + digits[3:12]
                        if digits.startswith('0') and len(digits) >= 10:
                            return '+212' + digits[1:10]
                        return ''
                    except Exception:
                        return ''
                candidate.phone = _normalize_moroccan_phone_edit(phone)
                candidate.address = address
                candidate.gender = gender
                candidate.linkedin_profile = linkedin_profile
                if birth_date:
                    candidate.birth_date = birth_date
                if photo:
                    candidate.photo = photo
                candidate.save()
                messages.success(request, f'Candidat {first_name} {last_name} modifié avec succès!')
                return redirect('recruitment_candidates')
            elif action == 'delete':
                candidate_id = request.POST.get('candidate_id')
                if not candidate_id:
                    messages.error(request, 'ID du candidat manquant.')
                    return redirect('recruitment_candidates')
                try:
                    candidate = Candidate.objects.get(id=candidate_id)
                    candidate_name = f"{candidate.first_name} {candidate.last_name}"
                    candidate.delete()
                    messages.success(request, f'Candidat {candidate_name} supprimé avec succès!')
                except Candidate.DoesNotExist:
                    messages.error(request, 'Candidat introuvable.')
                return redirect('recruitment_candidates')
            elif action == 'unarchive':
                candidate_id = request.POST.get('candidate_id')
                if not candidate_id:
                    messages.error(request, 'ID du candidat manquant.')
                    return redirect('recruitment_candidates')
                try:
                    candidate = Candidate.objects.get(id=candidate_id)
                    candidate.archived = False
                    candidate.save()
                    messages.success(request, f'Candidat {candidate.first_name} {candidate.last_name} réactivé avec succès!')
                except Candidate.DoesNotExist:
                    messages.error(request, 'Candidat introuvable.')
                return redirect('recruitment_candidates')
        except Exception as e:
            messages.error(request, f'Erreur lors du traitement du candidat: {str(e)}')
            return redirect('recruitment_candidates')

    # GET request - affichage de la liste et recherche
    candidates = Candidate.objects.filter(archived=False).order_by('last_name', 'first_name')
    archived_candidates = Candidate.objects.filter(archived=True).order_by('last_name', 'first_name')
    
    # Recherche pour les candidats actifs
    query = request.GET.get('q', '').strip()
    if query:
        candidates = candidates.filter(Q(last_name__icontains=query) | Q(first_name__icontains=query))
    
    # Recherche pour les candidats archivés
    archived_candidate_query = request.GET.get('archived_candidate_q', '').strip()
    if archived_candidate_query:
        archived_candidates = archived_candidates.filter(Q(last_name__icontains=archived_candidate_query) | Q(first_name__icontains=archived_candidate_query))
    
    # Ajout pagination
    from django.core.paginator import Paginator
    paginator = Paginator(candidates, 5)  # 5 candidats par page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    context = {
        'page_obj': page_obj,
        'query': query,
        'archived_candidates': archived_candidates,
        'archived_candidate_query': archived_candidate_query,
    }
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        from django.template.loader import render_to_string
        import re
        html_content = render_to_string('candidate.html', context)
        
        # Vérifier si c'est une recherche pour les candidats archivés
        if 'archived_candidate_q' in request.GET:
            # Extraire seulement la partie du tableau des candidats archivés
            table_match = re.search(r'<tbody id="archived-candidates-table">(.*?)</tbody>', html_content, re.DOTALL)
            if table_match:
                table_content = table_match.group(1)
                return HttpResponse(table_content)
            else:
                return HttpResponse('<tr><td colspan="7" class="text-center text-muted">Aucun candidat archivé trouvé.</td></tr>')
        else:
            # Extraire seulement la partie du tableau des candidats actifs
            table_match = re.search(r'<tbody id="candidates-table">(.*?)</tbody>', html_content, re.DOTALL)
            if table_match:
                table_content = table_match.group(1)
                return HttpResponse(table_content)
            else:
                return HttpResponse('<tr><td colspan="12" class="text-center text-danger">Erreur: Impossible de trouver le tableau</td></tr>')
    return render(request, 'candidate.html', context)


def recruitment_applications(request):
    from .models import Application, Offer, Candidate
    from django.utils import timezone
    from django.core.paginator import Paginator
    archived_param = request.GET.get('archived')
    show_archived = archived_param == '1'
    
    if request.method == 'POST':
        action = request.POST.get('action', 'add')
        try:
            if action == 'add':
                offer_id = request.POST.get('offer')
                candidate_id = request.POST.get('candidate')
                submission_date = request.POST.get('submission_date')
                source = request.POST.get('source')
                source_other = request.POST.get('source_other')
                always_available = request.POST.get('always_available') == 'on'
                availability_start = request.POST.get('availability_start')
                availability_end = request.POST.get('availability_end')
                cv = request.FILES.get('cv')
                cover_letter = request.FILES.get('cover_letter')
                certificates = request.FILES.get('certificates')
                status = request.POST.get('status')
                
                if not offer_id or not candidate_id or not submission_date or not status:
                    messages.error(request, 'Veuillez remplir tous les champs obligatoires.')
                    return redirect('recruitment_applications')
                
                try:
                    offer = Offer.objects.get(id=offer_id)
                    candidate = Candidate.objects.get(id=candidate_id)
                except (Offer.DoesNotExist, Candidate.DoesNotExist):
                    messages.error(request, 'Offre ou candidat introuvable.')
                    return redirect('recruitment_applications')
                
                # Conversion de la date
                try:
                    submission_datetime = datetime.strptime(submission_date, '%Y-%m-%dT%H:%M')
                except ValueError:
                    messages.error(request, 'Format de date invalide.')
                    return redirect('recruitment_applications')
                
                # Validation des dates de disponibilité
                if not always_available:
                    # Si on n'est pas toujours disponible, on doit avoir au moins une date de début
                    if not availability_start:
                        messages.error(request, 'Veuillez saisir une date de début de disponibilité ou cocher "Toujours disponible".')
                        return redirect('recruitment_applications')
                    # Conversion des dates de disponibilité en objets date
                    availability_start_date = None
                    availability_end_date = None
                    try:
                        if availability_start:
                            availability_start_date = datetime.strptime(availability_start, '%Y-%m-%d').date()
                        if availability_end:
                            availability_end_date = datetime.strptime(availability_end, '%Y-%m-%d').date()
                    except ValueError:
                        messages.error(request, 'Format de date de disponibilité invalide.')
                        return redirect('recruitment_applications')
                    # Si on a les deux dates, on valide l'intervalle
                    if availability_start_date and availability_end_date:
                        if availability_start_date > availability_end_date:
                            messages.error(request, 'La date de début de disponibilité ne peut pas être postérieure à la date de fin.')
                            return redirect('recruitment_applications')
                        if availability_start_date < offer.publication_date:
                            messages.error(request, f'La date de début de disponibilité ne peut pas être antérieure à la date de début de l\'offre ({offer.publication_date.strftime("%d/%m/%Y")}).')
                            return redirect('recruitment_applications')
                    # Si on n'a que la date de début, on valide juste qu'elle n'est pas avant l'offre
                    elif availability_start_date:
                        if availability_start_date < offer.publication_date:
                            messages.error(request, f'La date de début de disponibilité ne peut pas être antérieure à la date de début de l\'offre ({offer.publication_date.strftime("%d/%m/%Y")}).')
                            return redirect('recruitment_applications')
                
                # Traitement de la source
                final_source = source
                if source == 'other' and source_other:
                    final_source = source_other
                
                application = Application.objects.create(
                    offer=offer,
                    candidate=candidate,
                    submission_date=submission_datetime,
                    source=final_source,
                    always_available=always_available,
                    status=status
                )
                
                # Traitement des dates de disponibilité seulement si elles sont renseignées et pas toujours disponible
                if not always_available:
                    if availability_start:
                        application.availability_start = availability_start_date
                    if availability_end:
                        application.availability_end = availability_end_date
                
                if cv:
                    application.cv = cv
                if cover_letter:
                    application.cover_letter = cover_letter
                if certificates:
                    application.certificates = certificates
                application.save()
                
                messages.success(request, f'Candidature de {candidate.first_name} {candidate.last_name} pour {offer.title} ajoutée avec succès!')
                return redirect('recruitment_applications')
                
            elif action == 'edit':
                application_id = request.POST.get('application_id')
                offer_id = request.POST.get('offer')
                candidate_id = request.POST.get('candidate')
                submission_date = request.POST.get('submission_date')
                source = request.POST.get('source')
                source_other = request.POST.get('source_other')
                always_available = request.POST.get('always_available') == 'on'
                availability_start = request.POST.get('availability_start')
                availability_end = request.POST.get('availability_end')
                cv = request.FILES.get('cv')
                cover_letter = request.FILES.get('cover_letter')
                certificates = request.FILES.get('certificates')
                status = request.POST.get('status')
                
                if not application_id or not offer_id or not candidate_id or not submission_date or not status:
                    messages.error(request, 'Veuillez remplir tous les champs obligatoires.')
                    return redirect('recruitment_applications')
                
                try:
                    application = Application.objects.get(id=application_id)
                    offer = Offer.objects.get(id=offer_id)
                    candidate = Candidate.objects.get(id=candidate_id)
                except (Application.DoesNotExist, Offer.DoesNotExist, Candidate.DoesNotExist):
                    messages.error(request, 'Candidature, offre ou candidat introuvable.')
                    return redirect('recruitment_applications')
                
                # Conversion de la date
                try:
                    submission_datetime = datetime.strptime(submission_date, '%Y-%m-%dT%H:%M')
                except ValueError:
                    messages.error(request, 'Format de date invalide.')
                    return redirect('recruitment_applications')
                
                # Validation des dates de disponibilité
                availability_start_date = None
                availability_end_date = None
                if not always_available and (availability_start or availability_end):
                    try:
                        if availability_start:
                            availability_start_date = datetime.strptime(availability_start, '%Y-%m-%d').date()
                        if availability_end:
                            availability_end_date = datetime.strptime(availability_end, '%Y-%m-%d').date()
                    except ValueError:
                        messages.error(request, 'Format de date de disponibilité invalide.')
                        return redirect('recruitment_applications')
                    if availability_start_date and availability_end_date:
                        if availability_start_date > availability_end_date:
                            messages.error(request, 'La date de début de disponibilité ne peut pas être postérieure à la date de fin.')
                            return redirect('recruitment_applications')
                        if availability_start_date < offer.publication_date:
                            messages.error(request, f'La date de début de disponibilité ne peut pas être antérieure à la date de début de l\'offre ({offer.publication_date.strftime("%d/%m/%Y")}).')
                            return redirect('recruitment_applications')
                
                # Traitement de la source
                final_source = source
                if source == 'other' and source_other:
                    final_source = source_other
                
                application.offer = offer
                application.candidate = candidate
                application.submission_date = submission_datetime
                application.source = final_source
                application.always_available = always_available
                application.status = status
                
                # Traitement des dates de disponibilité seulement si elles sont renseignées et pas toujours disponible
                if not always_available:
                    if availability_start:
                        application.availability_start = availability_start_date
                    if availability_end:
                        application.availability_end = availability_end_date
                else:
                    # Si toujours disponible, on efface les dates
                    application.availability_start = None
                    application.availability_end = None
                
                if cv:
                    application.cv = cv
                if cover_letter:
                    application.cover_letter = cover_letter
                if certificates:
                    application.certificates = certificates
                application.save()
                
                messages.success(request, f'Candidature de {candidate.first_name} {candidate.last_name} pour {offer.title} modifiée avec succès!')
                return redirect('recruitment_applications')
                
            elif action == 'delete':
                application_id = request.POST.get('application_id')
                
                if not application_id:
                    messages.error(request, 'ID de la candidature manquant.')
                    return redirect('recruitment_applications')
                
                try:
                    application = Application.objects.get(id=application_id)
                    candidate_name = f"{application.candidate.first_name} {application.candidate.last_name}"
                    offer_title = application.offer.title
                    application.delete()
                    messages.success(request, f'Candidature de {candidate_name} pour {offer_title} supprimée avec succès!')
                except Application.DoesNotExist:
                    messages.error(request, 'Candidature introuvable.')
                
                return redirect('recruitment_applications')
                
        except Exception as e:
            messages.error(request, f'Erreur lors du traitement de la candidature: {str(e)}')
            return redirect('recruitment_applications')

    # GET request - affichage de la liste et recherche
    query = request.GET.get('q', '').strip()
    applications = Application.objects.select_related('offer', 'candidate').all()
    
    # Recherche pour les candidatures actives
    if query:
        applications = applications.filter(
            Q(candidate__first_name__icontains=query) |
            Q(candidate__last_name__icontains=query) |
            Q(offer__title__icontains=query)
        )
    if show_archived:
        applications = applications.filter(archived=True)
    else:
        applications = applications.filter(archived=False)
    applications = applications.order_by('-submission_date')
    
    # Recherche pour les candidatures archivées
    archived_application_query = request.GET.get('archived_application_q', '').strip()
    archived_applications = Application.objects.select_related('offer', 'candidate').filter(archived=True).order_by('-submission_date')
    if archived_application_query:
        archived_applications = archived_applications.filter(
            Q(candidate__first_name__icontains=archived_application_query) |
            Q(candidate__last_name__icontains=archived_application_query) |
            Q(offer__title__icontains=archived_application_query)
        )
    paginator = Paginator(applications, 5)  # 5 candidats par page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    offers = Offer.objects.filter(archived=False, end_date__gte=timezone.now().date()).order_by('title')
    candidates = Candidate.objects.filter(archived=False).order_by('first_name', 'last_name')
    context = {
        'page_obj': page_obj,
        'offers': offers,
        'candidates': candidates,
        'query': query,
        'show_archived': show_archived,
        'archived_applications': archived_applications,
        'archived_application_query': archived_application_query,
    }
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        from django.template.loader import render_to_string
        import re
        html_content = render_to_string('application.html', context)
        # Extraire le tbody
        table_match = re.search(r'<tbody id="applications-table">(.*?)</tbody>', html_content, re.DOTALL)
        tbody_html = table_match.group(1) if table_match else ''
        # Extraire la pagination
        pag_match = re.search(r'<div class="d-flex justify-content-end mt-2" id="pagination-wrapper">(.*?)</div>', html_content, re.DOTALL)
        pag_html = pag_match.group(1) if pag_match else ''
        # Retourner les deux blocs pour le JS
        from django.http import JsonResponse
        return JsonResponse({'tbody': tbody_html, 'pagination': pag_html})
    return render(request, 'application.html', context)


def recruitment_interviews(request):
    from .models import Interview, Application
    from django.utils import timezone
    if request.method == 'POST':
        action = request.POST.get('action', 'add')
        try:
            if action == 'add':
                application_id = request.POST.get('application')
                start = request.POST.get('start')
                end = request.POST.get('end')
                duration = request.POST.get('duration')
                location = request.POST.get('location')
                videcall_url = request.POST.get('videcall_url')
                organizer_id = request.POST.get('organizer')
                organizer = Employee.objects.get(id=organizer_id) if organizer_id else None
                description = request.POST.get('description')
                notes = request.POST.get('notes')
                result = request.POST.get('result')
                if not application_id or not start or not end:
                    messages.error(request, 'Veuillez remplir tous les champs obligatoires.')
                    return redirect('recruitment_interviews')
                try:
                    application = Application.objects.get(id=application_id)
                except Application.DoesNotExist:
                    messages.error(request, 'Candidature introuvable.')
                    return redirect('recruitment_interviews')
                # Vérification blocage entretien
                from .models import Interview
                previous_interviews = Interview.objects.filter(application=application).order_by('-start')
                if previous_interviews.exists():
                    last_result = previous_interviews.first().result
                    if last_result != 'positive':
                        messages.error(request, "Le candidat ne peut passer un nouvel entretien que si le résultat du précédent est positif.")
                        return redirect('recruitment_interviews')
                try:
                    start_dt = datetime.strptime(start, '%Y-%m-%dT%H:%M')
                    end_dt = datetime.strptime(end, '%Y-%m-%dT%H:%M')
                except ValueError:
                    messages.error(request, 'Format de date invalide.')
                    return redirect('recruitment_interviews')
                
                # Validation des dates d'entretien par rapport à la disponibilité de la candidature
                if not application.always_available:
                    # Vérifier que la date de début de l'entretien est dans l'intervalle de disponibilité
                    if application.availability_start:
                        if start_dt.date() < application.availability_start:
                            messages.error(request, f'La date de début de l\'entretien ne peut pas être antérieure à la date de début de disponibilité ({application.availability_start.strftime("%d/%m/%Y")}).')
                            return redirect('recruitment_interviews')
                    
                    if application.availability_end:
                        if start_dt.date() > application.availability_end:
                            messages.error(request, f'La date de début de l\'entretien ne peut pas être postérieure à la date de fin de disponibilité ({application.availability_end.strftime("%d/%m/%Y")}).')
                            return redirect('recruitment_interviews')
                    
                    # Vérifier que la date de fin de l'entretien est dans l'intervalle de disponibilité
                    if application.availability_start and end_dt.date() < application.availability_start:
                        messages.error(request, f'La date de fin de l\'entretien ne peut pas être antérieure à la date de début de disponibilité ({application.availability_start.strftime("%d/%m/%Y")}).')
                        return redirect('recruitment_interviews')
                    
                    if application.availability_end and end_dt.date() > application.availability_end:
                        messages.error(request, f'La date de fin de l\'entretien ne peut pas être postérieure à la date de fin de disponibilité ({application.availability_end.strftime("%d/%m/%Y")}).')
                        return redirect('recruitment_interviews')
                from datetime import timedelta
                try:
                    h, m = map(int, duration.split(':')) if duration else (0, 0)
                    duration_td = timedelta(hours=h, minutes=m)
                except Exception:
                    duration_td = None
                interview = Interview.objects.create(
                    application=application,
                    start=start_dt,
                    end=end_dt,
                    duration=duration_td,
                    location=location,
                    videcall_url=videcall_url,
                    organizer=organizer,
                    description=description,
                    notes=notes,
                    result=result,
                    archived=(result == 'positive')
                )
                messages.success(request, 'Entretien ajouté avec succès!')
                return redirect('recruitment_interviews')
            elif action == 'edit':
                interview_id = request.POST.get('interview_id')
                if not interview_id:
                    messages.error(request, 'ID de l\'entretien manquant.')
                    return redirect('recruitment_interviews')
                try:
                    interview = Interview.objects.get(id=interview_id)
                except Interview.DoesNotExist:
                    messages.error(request, 'Entretien introuvable.')
                    return redirect('recruitment_interviews')

                application_id = request.POST.get('application')
                start = request.POST.get('start')
                end = request.POST.get('end')
                duration = request.POST.get('duration')
                location = request.POST.get('location')
                videcall_url = request.POST.get('videcall_url')
                organizer_id = request.POST.get('organizer')
                organizer = Employee.objects.get(id=organizer_id) if organizer_id else None
                description = request.POST.get('description')
                notes = request.POST.get('notes')
                result = request.POST.get('result')

                if not application_id or not start or not end:
                    messages.error(request, 'Veuillez remplir tous les champs obligatoires.')
                    return redirect('recruitment_interviews')

                try:
                    application = Application.objects.get(id=application_id)
                    interview.application = application
                except Application.DoesNotExist:
                    messages.error(request, 'Candidature introuvable.')
                    return redirect('recruitment_interviews')

                try:
                    start_dt = datetime.strptime(start, '%Y-%m-%dT%H:%M')
                    end_dt = datetime.strptime(end, '%Y-%m-%dT%H:%M')
                    interview.start = start_dt
                    interview.end = end_dt
                except ValueError:
                    messages.error(request, 'Format de date invalide.')
                    return redirect('recruitment_interviews')
                
                # Validation des dates d'entretien par rapport à la disponibilité de la candidature
                if not application.always_available:
                    # Vérifier que la date de début de l'entretien est dans l'intervalle de disponibilité
                    if application.availability_start:
                        if start_dt.date() < application.availability_start:
                            messages.error(request, f'La date de début de l\'entretien ne peut pas être antérieure à la date de début de disponibilité ({application.availability_start.strftime("%d/%m/%Y")}).')
                            return redirect('recruitment_interviews')
                    
                    if application.availability_end:
                        if start_dt.date() > application.availability_end:
                            messages.error(request, f'La date de début de l\'entretien ne peut pas être postérieure à la date de fin de disponibilité ({application.availability_end.strftime("%d/%m/%Y")}).')
                            return redirect('recruitment_interviews')
                    
                    # Vérifier que la date de fin de l'entretien est dans l'intervalle de disponibilité
                    if application.availability_start and end_dt.date() < application.availability_start:
                        messages.error(request, f'La date de fin de l\'entretien ne peut pas être antérieure à la date de début de disponibilité ({application.availability_start.strftime("%d/%m/%Y")}).')
                        return redirect('recruitment_interviews')
                    
                    if application.availability_end and end_dt.date() > application.availability_end:
                        messages.error(request, f'La date de fin de l\'entretien ne peut pas être postérieure à la date de fin de disponibilité ({application.availability_end.strftime("%d/%m/%Y")}).')
                        return redirect('recruitment_interviews')

                from datetime import timedelta
                if duration:
                    try:
                        h, m = map(int, duration.split(':'))
                        interview.duration = timedelta(hours=h, minutes=m)
                    except Exception:
                        pass

                interview.location = location
                interview.videcall_url = videcall_url
                interview.organizer = organizer
                interview.description = description
                interview.notes = notes
                interview.result = result
                interview.archived = (result == 'positive')
                interview.save()
                messages.success(request, 'Entretien modifié avec succès!')
                return redirect('recruitment_interviews')
            elif action == 'delete':
                interview_id = request.POST.get('interview_id')
                if not interview_id:
                    messages.error(request, 'ID de l\'entretien manquant.')
                    return redirect('recruitment_interviews')
                try:
                    interview = Interview.objects.get(id=interview_id)
                    interview.delete()
                    messages.success(request, 'Entretien supprimé avec succès!')
                except Interview.DoesNotExist:
                    messages.error(request, 'Entretien introuvable.')
                return redirect('recruitment_interviews')
        except Exception as e:
            messages.error(request, f'Erreur lors du traitement de l\'entretien: {str(e)}')
            return redirect('recruitment_interviews')
    # Filtrer pour n'afficher que les entretiens non archivés
    interviews = Interview.objects.select_related('application', 'application__candidate', 'application__offer').filter(archived=False).order_by('-start')
    query = request.GET.get('q', '').strip()
    if query:
        interviews = Interview.objects.select_related('application', 'application__candidate', 'application__offer', 'organizer').filter(
            Q(location__icontains=query) |
            Q(application__offer__title__icontains=query) |
            Q(application__candidate__first_name__icontains=query) |
            Q(application__candidate__last_name__icontains=query) |
            Q(organizer__full_name__icontains=query),
            archived=False
        ).order_by('-start')
    
    # Récupérer les entretiens archivés pour le modal
    archived_interviews = Interview.objects.select_related('application', 'application__candidate', 'application__offer', 'organizer')\
        .filter(archived=True).order_by('-start')
    
    # Recherche pour les entretiens archivés
    archived_interview_query = request.GET.get('archived_interview_q', '').strip()
    if archived_interview_query:
        archived_interviews = archived_interviews.filter(
            Q(application__candidate__first_name__icontains=archived_interview_query) |
            Q(application__candidate__last_name__icontains=archived_interview_query) |
            Q(application__offer__title__icontains=archived_interview_query)
        )
    # Seules les candidatures acceptées et non archivées doivent être proposées pour un entretien
    applications = Application.objects.filter(status='accepted', archived=False).order_by('-submission_date')
    employees = Employee.objects.all().order_by('full_name')
    # Préparation des données Kanban
    kanban_data = {
        'reception': Interview.objects.select_related('application', 'application__candidate', 'application__offer').filter(stage='reception', archived=False).order_by('-start'),
        'a': Interview.objects.select_related('application', 'application__candidate', 'application__offer').filter(stage='a', archived=False).order_by('-start'),
        'b': Interview.objects.select_related('application', 'application__candidate', 'application__offer').filter(stage='b', archived=False).order_by('-start'),
    }
    paginator = Paginator(interviews, 5)  # 5 entretiens par page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    context = {
        'interviews': interviews,
        'page_obj': page_obj,
        'applications': applications,
        'employees': employees,
        'query': query,
        'kanban_data': kanban_data,
        'archived_interviews': archived_interviews,
        'archived_interview_query': archived_interview_query,
    }
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        from django.http import HttpResponse
        html_content = render_to_string('interview.html', context)
        import re
        
        # Vérifier si c'est une recherche pour les entretiens archivés
        if 'archived_interview_q' in request.GET:
            # Extraire seulement la partie du tableau des entretiens archivés
            table_match = re.search(r'<tbody id="archived-interviews-table">(.*?)</tbody>', html_content, re.DOTALL)
            if table_match:
                table_content = table_match.group(1)
                return HttpResponse(table_content)
            else:
                return HttpResponse('<tr><td colspan="8" class="text-center text-muted">Aucun entretien archivé trouvé.</td></tr>')
        else:
            # Extraire seulement la partie du tableau des entretiens actifs
            table_match = re.search(r'<tbody id="interviews-table">(.*?)</tbody>', html_content, re.DOTALL)
            if table_match:
                table_content = table_match.group(1)
                return HttpResponse(table_content)
            else:
                return HttpResponse('<tr><td colspan="10" class="text-center text-danger">Erreur: Impossible de trouver le tableau</td></tr>')
    # Après la pagination et la construction du context
    # On ajoute une propriété temporaire pour l'affichage du résultat dans le tableau
    for interview in page_obj:
        if interview.stage in ['reception', 'a']:
            interview.display_result = 'pending'
        else:
            interview.display_result = interview.result

    return render(request, 'interview.html', context)


def get_interview_json(request, interview_id):
    from .models import Interview
    try:
        interview = Interview.objects.get(id=interview_id)
        print(f"Récupération des données pour l'entretien {interview_id}")
        
        # Calculer la durée correctement
        duration_str = ''
        if interview.duration:
            total_seconds = interview.duration.total_seconds()
            hours = int(total_seconds // 3600)
            minutes = int((total_seconds % 3600) // 60)
            duration_str = f'{hours:02}:{minutes:02}'
        
        data = {
            'id': interview.id,
            'application': interview.application.id if interview.application else '',
            'start': interview.start.strftime('%Y-%m-%dT%H:%M') if interview.start else '',
            'end': interview.end.strftime('%Y-%m-%dT%H:%M') if interview.end else '',
            'duration': duration_str,
            'location': interview.location or '',
            'videcall_url': interview.videcall_url or '',
            'organizer': interview.organizer.full_name if interview.organizer else '',
            'description': interview.description or '',
            'notes': interview.notes or '',
            'result': interview.result or '',
        }
        print(f"Données JSON générées: {data}")
        return JsonResponse(data)
    except Interview.DoesNotExist:
        print(f"Entretien {interview_id} introuvable")
        return JsonResponse({'error': 'Entretien introuvable.'}, status=404)
    except Exception as e:
        print(f"Erreur lors de la récupération de l'entretien {interview_id}: {str(e)}")
        return JsonResponse({'error': f'Erreur: {str(e)}'}, status=500)


def recruitment_statistics(request):
    from django.db.models import Count, Q
    from django.utils import timezone
    from datetime import timedelta, datetime
    from calendar import monthrange
    
    # Périodes pour les filtres
    today = timezone.now().date()
    start_of_week = today - timedelta(days=today.weekday())
    end_of_week = start_of_week + timedelta(days=6)
    start_of_month = today.replace(day=1)
    end_of_month = today.replace(day=monthrange(today.year, today.month)[1])
    start_of_year = today.replace(month=1, day=1)
    end_of_year = today.replace(month=12, day=31)
    
    # Récupérer la période demandée
    period = request.GET.get('period', 'all')
    
    # Filtres de date selon la période
    if period == 'week':
        start_date = start_of_week
        end_date = end_of_week
    elif period == 'month':
        start_date = start_of_month
        end_date = end_of_month
    elif period == 'year':
        start_date = start_of_year
        end_date = end_of_year
    else:
        start_date = None
        end_date = None
    
    # Statistiques des candidats
    candidates_query = Candidate.objects.all()
    if start_date and end_date:
        candidates_query = candidates_query.filter(
            Q(archived=False) | Q(archived=True, id__in=Application.objects.filter(
                submission_date__date__range=[start_date, end_date]
            ).values('candidate'))
        )
    
    total_candidates = candidates_query.count()
    active_candidates = candidates_query.filter(archived=False).count()
    archived_candidates = candidates_query.filter(archived=True).count()
    
    # Statistiques des candidatures
    applications_query = Application.objects.all()
    if start_date and end_date:
        applications_query = applications_query.filter(submission_date__date__range=[start_date, end_date])
    
    total_applications = applications_query.count()
    applications_by_status = applications_query.values('status').annotate(count=Count('id'))
    
    # Statistiques des entretiens
    interviews_query = Interview.objects.all()
    if start_date and end_date:
        interviews_query = interviews_query.filter(start__date__range=[start_date, end_date])
    
    total_interviews = interviews_query.count()
    interviews_by_result = interviews_query.values('result').annotate(count=Count('id'))
    interviews_by_stage = interviews_query.values('stage').annotate(count=Count('id'))
    
    # Statistiques des offres
    offers_query = Offer.objects.all()
    if start_date and end_date:
        offers_query = offers_query.filter(publication_date__date__range=[start_date, end_date])
    
    total_offers = offers_query.count()
    active_offers = offers_query.filter(archived=False, end_date__gte=today).count()
    archived_offers = offers_query.filter(archived=True).count()
    
    # Statistiques par département
    department_stats = []
    if start_date and end_date:
        dept_offers = offers_query.filter(publication_date__date__range=[start_date, end_date])
    else:
        dept_offers = offers_query
    
    dept_counts = dept_offers.values('department__name').annotate(
        num_offers=Count('id'),
        num_applications=Count('application'),
        num_interviews=Count('application__interview')
    )
    
    for dept in dept_counts:
        department_stats.append({
            'name': dept['department__name'] or 'Sans département',
            'offers': dept['num_offers'],
            'applications': dept['num_applications'],
            'interviews': dept['num_interviews']
        })
    
    # Statistiques temporelles (derniers 12 mois)
    monthly_stats = []
    for i in range(12):
        month_date = today.replace(day=1) - timedelta(days=30*i)
        month_start = month_date.replace(day=1)
        month_end = month_date.replace(day=monthrange(month_date.year, month_date.month)[1])
        
        month_candidates = Candidate.objects.filter(
            Q(archived=False) | Q(archived=True, id__in=Application.objects.filter(
                submission_date__date__range=[month_start, month_end]
            ).values('candidate'))
        ).count()
        
        month_applications = Application.objects.filter(
            submission_date__date__range=[month_start, month_end]
        ).count()
        
        month_interviews = Interview.objects.filter(
            start__date__range=[month_start, month_end]
        ).count()
        
        monthly_stats.append({
            'month': month_date.strftime('%b %Y'),
            'candidates': month_candidates,
            'applications': month_applications,
            'interviews': month_interviews
        })
    
    monthly_stats.reverse()  # Du plus ancien au plus récent
    
    # Taux de conversion
    if total_applications > 0:
        conversion_rate = (total_interviews / total_applications) * 100
    else:
        conversion_rate = 0
    
    if total_interviews > 0:
        positive_interviews = interviews_query.filter(result='positive').count()
        success_rate = (positive_interviews / total_interviews) * 100
    else:
        success_rate = 0
    
    # Calcul des pourcentages pour les progress bars
    if total_candidates > 0:
        active_candidates_percentage = (active_candidates / total_candidates) * 100
        archived_candidates_percentage = (archived_candidates / total_candidates) * 100
    else:
        active_candidates_percentage = 0
        archived_candidates_percentage = 0
    
    if total_offers > 0:
        active_offers_percentage = (active_offers / total_offers) * 100
        archived_offers_percentage = (archived_offers / total_offers) * 100
    else:
        active_offers_percentage = 0
        archived_offers_percentage = 0
    
    # Calcul des pourcentages pour les tableaux
    applications_by_status_with_percentages = []
    for status in applications_by_status:
        percentage = (status['count'] / total_applications * 100) if total_applications > 0 else 0
        applications_by_status_with_percentages.append({
            'status': status['status'],
            'count': status['count'],
            'percentage': round(percentage, 1)
        })
    
    interviews_by_stage_with_percentages = []
    for stage in interviews_by_stage:
        percentage = (stage['count'] / total_interviews * 100) if total_interviews > 0 else 0
        interviews_by_stage_with_percentages.append({
            'stage': stage['stage'],
            'count': stage['count'],
            'percentage': round(percentage, 1)
        })
    
    context = {
        'period': period,
        'start_date': start_date,
        'end_date': end_date,
        'today': today,
        'total_candidates': total_candidates,
        'active_candidates': active_candidates,
        'archived_candidates': archived_candidates,
        'active_candidates_percentage': round(active_candidates_percentage, 1),
        'archived_candidates_percentage': round(archived_candidates_percentage, 1),
        'total_applications': total_applications,
        'applications_by_status': applications_by_status_with_percentages,
        'total_interviews': total_interviews,
        'interviews_by_result': interviews_by_result,
        'interviews_by_stage': interviews_by_stage_with_percentages,
        'total_offers': total_offers,
        'active_offers': active_offers,
        'archived_offers': archived_offers,
        'active_offers_percentage': round(active_offers_percentage, 1),
        'archived_offers_percentage': round(archived_offers_percentage, 1),
        'department_stats': department_stats,
        'monthly_stats': monthly_stats,
        'conversion_rate': round(conversion_rate, 1),
        'success_rate': round(success_rate, 1),
        'periods': [
            {'value': 'all', 'label': 'Toutes les périodes'},
            {'value': 'week', 'label': 'Cette semaine'},
            {'value': 'month', 'label': 'Ce mois'},
            {'value': 'year', 'label': 'Cette année'},
        ]
    }
    
    return render(request, 'statistics.html', context)


def recruitment_pointage(request):

    query = request.GET.get('q', '').strip()
    
    if request.method == 'POST' and request.FILES.get('csv_file'):
        csv_file = request.FILES['csv_file']
            
        # Vérifier que c'est bien un fichier CSV
        if not csv_file.name.endswith('.csv'):
            messages.error(request, 'Veuillez sélectionner un fichier CSV valide.')
            pointages = Pointage.objects.select_related('employee').all().order_by('-employee__employee_id')
            paginator = Paginator(pointages, 10)
            page_number = request.GET.get('page')
            page_obj = paginator.get_page(page_number)
            return render(request, 'pointage.html', {'page_obj': page_obj, 'query': query})
        try:
            # Lire le contenu du fichier CSV
            csv_content = csv_file.read().decode('utf-8')
            csv_reader = csv.DictReader(io.StringIO(csv_content))
            # Structure attendue
            expected_columns = ['Code Employé', 'Nom', 'Prénom', 'Date', 'Jour', "Heure d'arrivée", 'Heure départ', 'Heure travaillée']
            # Vérifier les colonnes
            if not all(col in csv_reader.fieldnames for col in expected_columns):
                messages.error(request, f'Structure du fichier CSV invalide. Colonnes attendues: {", ".join(expected_columns)}')
                pointages = Pointage.objects.select_related('employee').all().order_by('-employee__employee_id')
                paginator = Paginator(pointages, 10)
                page_number = request.GET.get('page')
                page_obj = paginator.get_page(page_number)
                return render(request, 'pointage.html', {'page_obj': page_obj, 'query': query})
            
            inserted_count = 0
            errors_count = 0
            errors_list = []
            
            for row in csv_reader:
                code_employe = row.get('Code Employé', '').strip()
                nom = row.get('Nom', '').strip()
                prenom = row.get('Prénom', '').strip()
                date_str = row.get('Date', '').strip()
                jour = row.get('Jour', '').strip()
                heure_arrivee = row.get("Heure d'arrivée", '').strip()
                heure_depart = row.get('Heure départ', '').strip()
                heure_travaillee = row.get('Heure travaillée', '').strip()
                
                from datetime import datetime
                try:
                    date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
                except Exception:
                    date_obj = None
                
                # Rechercher l'employé correspondant au code_employe
                try:
                    employee = Employee.objects.get(employee_id=code_employe)
                except Employee.DoesNotExist:
                    errors_count += 1
                    errors_list.append(f"Employé avec le code '{code_employe}' non trouvé")
                    continue
                
                # Conversion des heures
                try:
                    heure_arrivee_obj = datetime.strptime(heure_arrivee, '%H:%M').time()
                    heure_depart_obj = datetime.strptime(heure_depart, '%H:%M').time()
                except ValueError:
                    errors_count += 1
                    errors_list.append(f"Format d'heure invalide pour l'employé {code_employe}")
                    continue
                
                # Conversion de la durée travaillée
                try:
                    # Nettoyer la chaîne
                    heure_travaillee = heure_travaillee.strip()
                    print(f"Durée brute pour {code_employe}: '{heure_travaillee}'")
                    
                    if heure_travaillee and ':' in heure_travaillee:
                        # Format "HH:MM" ou "H:M"
                        parts = heure_travaillee.split(':')
                        if len(parts) >= 2:
                            hours = int(parts[0])
                            minutes = int(parts[1])
                            duration_obj = timedelta(hours=hours, minutes=minutes)
                        else:
                            print(f"Format de durée non reconnu pour {code_employe}: {heure_travaillee}")
                            duration_obj = timedelta()
                    elif heure_travaillee and 'h' in heure_travaillee:
                        # Format "9h00", "8h40", etc.
                        parts = heure_travaillee.split('h')
                        if len(parts) == 2:
                            hours = int(parts[0])
                            minutes = int(parts[1])
                            duration_obj = timedelta(hours=hours, minutes=minutes)
                        else:
                            print(f"Format de durée non reconnu pour {code_employe}: {heure_travaillee}")
                            duration_obj = timedelta()
                    elif heure_travaillee:
                        # Essayer de convertir en heures décimales (ex: "8.5" = 8h30)
                        try:
                            hours_decimal = float(heure_travaillee)
                            hours = int(hours_decimal)
                            minutes = int((hours_decimal - hours) * 60)
                            duration_obj = timedelta(hours=hours, minutes=minutes)
                        except ValueError:
                            print(f"Format de durée non reconnu pour {code_employe}: {heure_travaillee}")
                            duration_obj = timedelta()
                    else:
                        duration_obj = timedelta()
                    
                    print(f"Durée convertie pour {code_employe}: {duration_obj}")
                except Exception as e:
                    print(f"Erreur conversion durée pour {code_employe}: {str(e)}")
                    duration_obj = timedelta()
                
                # Ajout à la base de données si la date est valide
                if date_obj:
                    try:
                        Pointage.objects.create(
                            employee=employee,
                            date=date_obj,
                            jour=jour,
                            heure_arrivee=heure_arrivee_obj,
                            heure_depart=heure_depart_obj,
                            heure_travaillee=duration_obj
                        )
                        inserted_count += 1
                    except Exception as e:
                        errors_count += 1
                        errors_list.append(f"Erreur lors de l'enregistrement pour l'employé {code_employe}: {str(e)}")
                        print(f"Erreur lors de l'enregistrement en base : {e}")
                else:
                    errors_count += 1
                    errors_list.append(f"Date invalide pour l'employé {code_employe}")
            
            # Messages de résultat
            if inserted_count > 0:
                messages.success(request, f'Fichier CSV traité avec succès ! {inserted_count} lignes insérées en base.')
            if errors_count > 0:
                error_message = f'{errors_count} erreurs rencontrées. '
                if len(errors_list) <= 5:
                    error_message += 'Erreurs: ' + '; '.join(errors_list)
                else:
                    error_message += f'Premières erreurs: {"; ".join(errors_list[:5])}...'
                messages.warning(request, error_message)
            if inserted_count == 0 and errors_count == 0:
                messages.error(request, 'Aucune ligne insérée en base. Vérifiez le format du fichier CSV.')
                
        except UnicodeDecodeError:
            messages.error(request, 'Erreur de décodage du fichier. Assurez-vous que le fichier est encodé en UTF-8.')
        except Exception as e:
            messages.error(request, f'Erreur lors du traitement du fichier: {str(e)}')
    
    # Affichage de la liste paginée avec recherche
    pointages = Pointage.objects.select_related('employee').all()
    
    # Recherche
    if query:
        pointages = pointages.filter(
            Q(employee__employee_id__icontains=query) |
            Q(employee__full_name__icontains=query) |
            Q(employee__email__icontains=query)
        )
    
    pointages = pointages.order_by('-employee__employee_id')
    paginator = Paginator(pointages, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    return render(request, 'pointage.html', {'page_obj': page_obj, 'query': query})


@csrf_exempt
def recruitment_interview_kanban_stage(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            interview_id = data.get('id')
            new_stage = data.get('stage')
            if not interview_id or not new_stage:
                return JsonResponse({'success': False, 'error': 'Paramètres manquants.'})
            interview = Interview.objects.get(id=interview_id)
            interview.stage = new_stage
            # Mise à jour automatique du résultat selon la colonne
            if new_stage in ['reception', 'a']:
                interview.result = 'pending'
            elif new_stage == 'b':
                interview.result = None
            interview.save()
            return JsonResponse({'success': True})
        except Interview.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Entretien introuvable.'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    return JsonResponse({'success': False, 'error': 'Méthode non autorisée.'})


def export_candidates_excel(request):
    from .models import Candidate
    import datetime
    candidates = Candidate.objects.all().order_by('last_name', 'first_name')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Candidats'
    headers = ['Nom', 'Prénom', 'Email', 'Téléphone', 'Date de naissance', 'Adresse', 'Genre', 'Profil LinkedIn']
    ws.append(headers)
    for c in candidates:
        ws.append([
            c.last_name,
            c.first_name,
            c.email,
            c.phone or '',
            c.birth_date.strftime('%d/%m/%Y') if c.birth_date else '',
            c.address or '',
            c.gender or '',
            c.linkedin_profile or '',
        ])
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"candidats_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)
    return response

def export_candidates_pdf(request):
    from .models import Candidate
    from django.http import JsonResponse
    import json
    from datetime import datetime
    
    # Récupérer tous les candidats
    candidates = Candidate.objects.all().order_by('last_name', 'first_name')
    
    # Préparer les données pour le PDF
    candidates_data = []
    for candidate in candidates:
        candidates_data.append([
            candidate.last_name or '',
            candidate.first_name or '',
            candidate.email or '',
            candidate.phone or '',
            candidate.birth_date.strftime('%d/%m/%Y') if candidate.birth_date else '',
            candidate.address or '',
            candidate.gender or '',
            candidate.linkedin_profile or '',
        ])
    
    # Retourner les données au format JSON pour le JavaScript
    return JsonResponse({
        'title': 'Liste des Candidats',
        'headers': ['Nom', 'Prénom', 'Email', 'Téléphone', 'Date de naissance', 'Adresse', 'Genre', 'Profil LinkedIn'],
        'data': candidates_data,
        'filename': f'candidats_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
    })


def recruitment_configuration(request):
    from .models import Department, Employee
    from django.contrib import messages
    action = request.POST.get('action') if request.method == 'POST' else None
    tab = request.POST.get('tab', 'departments') if request.method == 'POST' else request.GET.get('tab', 'departments')
    # Recherche
    query = request.GET.get('q', '').strip()
    # CRUD Department
    if action in ['add_department', 'edit_department', 'delete_department']:
        if action == 'add_department':
            name = request.POST.get('name')
            description = request.POST.get('description')
            if not name:
                messages.error(request, 'Le nom du département est obligatoire.')
            else:
                Department.objects.create(name=name, description=description)
                messages.success(request, f'Département "{name}" ajouté avec succès!')
            tab = 'departments'
        elif action == 'edit_department':
            dep_id = request.POST.get('department_id')
            name = request.POST.get('name')
            description = request.POST.get('description')
            try:
                dep = Department.objects.get(id=dep_id)
                dep.name = name
                dep.description = description
                dep.save()
                messages.success(request, f'Département "{name}" modifié avec succès!')
            except Department.DoesNotExist:
                messages.error(request, 'Département introuvable.')
            tab = 'departments'
        elif action == 'delete_department':
            dep_id = request.POST.get('department_id')
            try:
                dep = Department.objects.get(id=dep_id)
                dep.delete()
                messages.success(request, 'Département supprimé avec succès!')
            except Department.DoesNotExist:
                messages.error(request, 'Département introuvable.')
            tab = 'departments'
    # CRUD Employee
    elif action in ['add_employee', 'edit_employee', 'delete_employee']:
        if action == 'add_employee':
            employee_id = request.POST.get('employee_id')
            full_name = request.POST.get('full_name')
            birth_date = request.POST.get('birth_date')
            birth_date = birth_date if birth_date else None
            photo = request.FILES.get('photo')
            marital_status = request.POST.get('marital_status')
            gender = request.POST.get('gender')
            email = request.POST.get('email')
            personal_phone = request.POST.get('personal_phone')
            work_phone = request.POST.get('work_phone')
            personal_address = request.POST.get('personal_address')
            work_address = request.POST.get('work_address')
            children_count = request.POST.get('children_count') or 0
            status = request.POST.get('status')
            department_id = request.POST.get('department')
            work_certificate = request.FILES.get('work_certificate')
            legalized_contract = request.FILES.get('legalized_contract')
            doctor_agreement = request.FILES.get('doctor_agreement')
            temporary_agreement = request.FILES.get('temporary_agreement')
            base_salary = request.POST.get('base_salary')
            career_evolution = request.POST.get('career_evolution')
            skills = request.POST.get('skills')
            is_supervisor = request.POST.get('is_supervisor', 0)
            try:
                department = Department.objects.get(id=department_id) if department_id else None
            except Department.DoesNotExist:
                department = None
            emp = Employee(
                employee_id=employee_id,
                full_name=full_name,
                birth_date=birth_date,
                marital_status=marital_status,
                gender=gender,
                email=email,
                personal_phone=personal_phone,
                work_phone=work_phone,
                personal_address=personal_address,
                work_address=work_address,
                children_count=children_count,
                status=status,
                department=department,
                base_salary=base_salary,
                career_evolution=career_evolution,
                skills=skills,
                is_supervisor=is_supervisor
            )
            if photo:
                emp.photo = photo
            if work_certificate:
                emp.work_certificate = work_certificate
            if legalized_contract:
                emp.legalized_contract = legalized_contract
            if doctor_agreement:
                emp.doctor_agreement = doctor_agreement
            if temporary_agreement:
                emp.temporary_agreement = temporary_agreement
            emp.save()
            messages.success(request, f'Employé "{full_name}" ajouté avec succès!')
            tab = 'employees'
        elif action == 'edit_employee':
            emp_id = request.POST.get('emp_id')
            try:
                emp = Employee.objects.get(id=emp_id)
                emp.employee_id = request.POST.get('employee_id')
                emp.full_name = request.POST.get('full_name')
                birth_date = request.POST.get('birth_date')
                emp.birth_date = birth_date if birth_date else None
                emp.marital_status = request.POST.get('marital_status')
                emp.gender = request.POST.get('gender')
                emp.email = request.POST.get('email')
                emp.personal_phone = request.POST.get('personal_phone')
                emp.work_phone = request.POST.get('work_phone')
                emp.personal_address = request.POST.get('personal_address')
                emp.work_address = request.POST.get('work_address')
                emp.children_count = request.POST.get('children_count') or 0
                emp.status = request.POST.get('status')
                department_id = request.POST.get('department')
                try:
                    emp.department = Department.objects.get(id=department_id) if department_id else None
                except Department.DoesNotExist:
                    emp.department = None
                emp.base_salary = request.POST.get('base_salary')
                emp.career_evolution = request.POST.get('career_evolution')
                emp.skills = request.POST.get('skills')
                emp.is_supervisor = request.POST.get('is_supervisor', 0)
                if request.FILES.get('photo'):
                    emp.photo = request.FILES.get('photo')
                if request.FILES.get('work_certificate'):
                    emp.work_certificate = request.FILES.get('work_certificate')
                if request.FILES.get('legalized_contract'):
                    emp.legalized_contract = request.FILES.get('legalized_contract')
                if request.FILES.get('doctor_agreement'):
                    emp.doctor_agreement = request.FILES.get('doctor_agreement')
                if request.FILES.get('temporary_agreement'):
                    emp.temporary_agreement = request.FILES.get('temporary_agreement')
                emp.save()
                messages.success(request, f'Employé "{emp.full_name}" modifié avec succès!')
            except Employee.DoesNotExist:
                messages.error(request, 'Employé introuvable.')
            tab = 'employees'
        elif action == 'delete_employee':
            emp_id = request.POST.get('emp_id')
            try:
                emp = Employee.objects.get(id=emp_id)
                emp.delete()
                messages.success(request, 'Employé supprimé avec succès!')
            except Employee.DoesNotExist:
                messages.error(request, 'Employé introuvable.')
            tab = 'employees'
    # Filtrage pour AJAX ou affichage normal
    departments = Department.objects.all().order_by('name')
    employees = Employee.objects.select_related('department').all().order_by('full_name')
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        # Filtrage selon le tab et la recherche
        if tab == 'departments':
            if query:
                departments = departments.filter(name__icontains=query)
        elif tab == 'employees':
            if query:
                employees = employees.filter(
                    Q(employee_id__icontains=query) |
                    Q(full_name__icontains=query) |
                    Q(email__icontains=query) |
                    Q(personal_phone__icontains=query) |
                    Q(work_phone__icontains=query)
                )
        # Rendu du template complet
        context = {
            'departments': departments,
            'employees': employees,
            'tab': tab,
            'marital_status_choices': Employee.MARITAL_STATUS_CHOICES,
            'gender_choices': Employee.GENDER_CHOICES,
            'status_choices': Employee.STATUS_CHOICES,
        }
        from django.template.loader import render_to_string
        html_content = render_to_string('recruitment_configuration.html', context, request=request)
        import re
        if tab == 'departments':
            match = re.search(r'<tbody id="departments-table">(.*?)</tbody>', html_content, re.DOTALL)
            if match:
                return HttpResponse(match.group(1))
            else:
                return HttpResponse('<tr><td colspan="3" class="text-center text-danger">Erreur: Impossible de trouver le tableau</td></tr>')
        elif tab == 'employees':
            match = re.search(r'<tbody id="employees-table">(.*?)</tbody>', html_content, re.DOTALL)
            if match:
                return HttpResponse(match.group(1))
            else:
                return HttpResponse('<tr><td colspan="9" class="text-center text-danger">Erreur: Impossible de trouver le tableau</td></tr>')
    # Affichage normal
    context = {
        'departments': departments,
        'employees': employees,
        'tab': tab,
        'marital_status_choices': Employee.MARITAL_STATUS_CHOICES,
        'gender_choices': Employee.GENDER_CHOICES,
        'status_choices': Employee.STATUS_CHOICES,
    }
    return render(request, 'recruitment_configuration.html', context)

def interview_events_json(request):
    """Retourne les entretiens au format JSON pour le calendrier moderne."""
    from .models import Interview
    from datetime import datetime
    
    # Récupérer les paramètres de date
    start_date = request.GET.get('start')
    end_date = request.GET.get('end')
    
    # Filtrer les entretiens selon la période demandée
    interviews = Interview.objects.select_related('application', 'application__candidate', 'application__offer').filter(archived=False)
    
    if start_date:
        try:
            start_dt = datetime.strptime(start_date, '%Y-%m-%d')
            interviews = interviews.filter(start__date__gte=start_dt.date())
        except ValueError:
            pass
    
    if end_date:
        try:
            end_dt = datetime.strptime(end_date, '%Y-%m-%d')
            interviews = interviews.filter(start__date__lte=end_dt.date())
        except ValueError:
            pass
    
    events = []
    for interview in interviews:
        title = f"{interview.application.candidate} - {interview.application.offer}" if interview.application else "Entretien"
        events.append({
            'id': interview.id,
            'title': title,
            'start': interview.start.isoformat() if interview.start else '',
            'end': interview.end.isoformat() if interview.end else '',
            'location': interview.location or '',
            'organizer': interview.organizer.full_name if interview.organizer else '',
            'description': interview.description or '',
            'result': interview.result or '',
            'stage': interview.stage or '',
        })
    return JsonResponse(events, safe=False)


@csrf_exempt
def recruitment_interview_set_result(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            interview_id = data.get('id')
            result = data.get('result')
            # On accepte maintenant 'positive', 'negative', 'pending'
            if not interview_id or result not in ['positive', 'negative', 'pending']:
                return JsonResponse({'success': False, 'error': 'Paramètres manquants ou invalides.'})
            interview = Interview.objects.get(id=interview_id)
            application = interview.application
            offer = application.offer
            # Limite d'acceptation : vérifier le nombre d'entretiens déjà validés pour cette offre
            if result == 'positive':
                accepted_count = Interview.objects.filter(application__offer=offer, result='positive').count()
                if accepted_count >= offer.positions_available:
                    return JsonResponse({'success': False, 'error': f"Le nombre maximum de personnes acceptées pour cette offre ({offer.positions_available}) est déjà atteint."})
            interview.result = result
            interview.archived = (result == 'positive')
            interview.save()

            # AJOUT AUTOMATIQUE DANS EMPLOYEE ET ARCHIVAGE SI RESULTAT POSITIF
            if result == 'positive':
                candidate = interview.application.candidate
                # Ajouter dans Employee si pas déjà présent
                if not Employee.objects.filter(email__iexact=candidate.email.strip()).exists():
                    Employee.objects.create(
                        full_name=f'{candidate.first_name} {candidate.last_name}',
                        email=candidate.email,
                        birth_date=candidate.birth_date or datetime.now().date(),
                        gender=candidate.gender[0].upper() if candidate.gender else 'M',
                        personal_address=candidate.address or '',
                        employee_id=f'EMP{candidate.id:05d}',
                        status='A',
                        base_salary=0,
                        department=None,
                        marital_status='S',
                        personal_phone=candidate.phone or '',
                        work_phone='',
                        children_count=0,
                        is_supervisor=0,
                    )
                candidate.archived = True
                candidate.save()
                # Archiver la candidature associée
                application = interview.application
                application.archived = True
                application.save()
                # Archiver l'offre si le nombre de candidats acceptés atteint la limite
                accepted_count = Interview.objects.filter(application__offer=offer, result='positive').count()
                if accepted_count >= offer.positions_available:
                    offer.archived = True
                    offer.save()
            else:  # 'negative' ou 'pending'
                candidate = interview.application.candidate
                Employee.objects.filter(email__iexact=candidate.email.strip()).delete()
                candidate.archived = False
                candidate.save()
            return JsonResponse({'success': True})
        except Interview.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Entretien introuvable.'})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    return JsonResponse({'success': False, 'error': 'Méthode non autorisée.'})


@require_GET
def recruitment_archived_interviews(request):
    from .models import Interview
    from django.template.loader import render_to_string
    from django.http import HttpResponse
    
    # Récupérer tous les entretiens archivés
    archived_interviews = Interview.objects.select_related('application', 'application__candidate', 'application__offer', 'organizer')\
        .filter(archived=True).order_by('-start')
    
    # Recherche pour les entretiens archivés
    archived_interview_query = request.GET.get('archived_interview_q', '').strip()
    if archived_interview_query:
        archived_interviews = archived_interviews.filter(
            Q(application__candidate__first_name__icontains=archived_interview_query) |
            Q(application__candidate__last_name__icontains=archived_interview_query) |
            Q(application__offer__title__icontains=archived_interview_query)
        )
    
    # Préparer les données pour le template
    context = {
        'archived_interviews': archived_interviews,
        'archived_interview_query': archived_interview_query,
    }
    
    # Si c'est une requête AJAX, retourner seulement le contenu du tableau
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        html_content = render_to_string('interview.html', context)
        import re
        table_match = re.search(r'<tbody id="archived-interviews-table">(.*?)</tbody>', html_content, re.DOTALL)
        if table_match:
            table_content = table_match.group(1)
            return HttpResponse(table_content)
        else:
            return HttpResponse('<tr><td colspan="8" class="text-center text-muted">Aucun entretien archivé trouvé.</td></tr>')
    
    # Sinon, retourner les données JSON pour la compatibilité
    data = []
    for interview in archived_interviews:
        data.append({
            'candidate': str(interview.application.candidate) if interview.application and interview.application.candidate else '',
            'offer': str(interview.application.offer) if interview.application and interview.application.offer else '',
            'date': interview.start.strftime('%d/%m/%Y') if interview.start else '',
            'heure': interview.start.strftime('%H:%M') if interview.start else '',
            'lieu': interview.location or '',
            'organisateur': interview.organizer.full_name if interview.organizer and hasattr(interview.organizer, 'full_name') else '',
            'description': interview.description or '',
            'notes': interview.notes or '',
        })
    return JsonResponse(data, safe=False)

@require_GET
def department_distribution_api(request):
    from .models import Department, Offer
    from django.utils import timezone
    from django.db.models import Count, Q
    import json
    period = request.GET.get('period', 'all')
    today = timezone.now().date()
    start_of_week = today - timedelta(days=today.weekday())
    end_of_week = start_of_week + timedelta(days=6)
    start_of_month = today.replace(day=1)
    end_of_month = today.replace(day=monthrange(today.year, today.month)[1])
    start_of_year = today.replace(month=1, day=1)
    end_of_year = today.replace(month=12, day=31)
    def get_department_distribution(start_date=None, end_date=None):
        depts = Department.objects.all()
        offers = Offer.objects.all()
        if start_date and end_date:
            offers = offers.filter(publication_date__gte=start_date, publication_date__lte=end_date)
        elif start_date:
            offers = offers.filter(publication_date__gte=start_date)
        dept_offer_counts = offers.values('department').annotate(num_offers=Count('id'))
        dept_offer_map = {d['department']: d['num_offers'] for d in dept_offer_counts}
        result = []
        for dept in depts:
            result.append({'name': dept.name, 'num_offers': dept_offer_map.get(dept.id, 0)})
        return result
    if period == 'week':
        data = get_department_distribution(start_of_week, end_of_week)
    elif period == 'month':
        data = get_department_distribution(start_of_month, end_of_month)
    elif period == 'year':
        data = get_department_distribution(start_of_year, end_of_year)
    else:
        data = get_department_distribution()
    return JsonResponse(data, safe=False)


def recruitment_suivi(request):
    """
    Vue pour la page de suivi des employés (retards, absences, heures travaillées)
    """
    from datetime import datetime, time, timedelta
    from django.db.models import Q, F, ExpressionWrapper, fields, Count, Case, When, Value, IntegerField
    from django.db.models.functions import ExtractHour, ExtractMinute, Extract
    
    # Récupération des paramètres de filtrage
    query = request.GET.get('q', '').strip()
    date_debut = request.GET.get('dateDebut', '')
    date_fin = request.GET.get('dateFin', '')
    status_filter = request.GET.get('statusFilter', '')
    
    # Base queryset
    pointages = Pointage.objects.select_related('employee').all()
    
    # Filtrage par recherche (nom ou code employé)
    if query:
        pointages = pointages.filter(
            Q(employee__full_name__icontains=query) |
            Q(employee__employee_id__icontains=query)
        )
    
    # Filtrage par date
    if date_debut:
        try:
            debut = datetime.strptime(date_debut, '%Y-%m-%d').date()
            pointages = pointages.filter(date__gte=debut)
        except ValueError:
            pass
    
    if date_fin:
        try:
            fin = datetime.strptime(date_fin, '%Y-%m-%d').date()
            pointages = pointages.filter(date__lte=fin)
        except ValueError:
            pass
    
    # Règles métier :
    # - Heure d'arrivée normale : 9h00
    # - Retard : après 9h10 (10 minutes de tolérance)
    # - Heures de travail normales : 8h (de 9h à 17h)
    # - Heures insuffisantes : moins de 8h travaillées
    
    heure_arrivee_normale_minutes = 9 * 60  # 9h00 = 540 minutes
    retard_minutes = (9 * 60) + 10  # 9h10 = 550 minutes
    heures_travail_normales_secondes = 8 * 3600  # 8h = 28800 secondes
    
    # Annoter les pointages avec les statuts
    pointages = pointages.annotate(
        heure_arrivee_minutes=ExpressionWrapper(
            ExtractHour('heure_arrivee') * 60 + ExtractMinute('heure_arrivee'),
            output_field=fields.IntegerField()
        ),
        heure_travaillee_secondes=ExpressionWrapper(
            Extract('heure_travaillee', 'epoch'),
            output_field=fields.IntegerField()
        )
    )
    
    # Filtrage par statut
    if status_filter:
        if status_filter == 'present':
            pointages = pointages.filter(
                heure_arrivee_minutes__lte=retard_minutes,
                heure_travaillee_secondes__gte=heures_travail_normales_secondes
            )
        elif status_filter == 'retard':
            pointages = pointages.filter(heure_arrivee_minutes__gt=retard_minutes)
        elif status_filter == 'insuffisant':
            pointages = pointages.filter(heure_travaillee_secondes__lt=heures_travail_normales_secondes)
        elif status_filter == 'absent':
            # Pour les absences, on cherche les employés qui n'ont pas de pointage
            # sur une période donnée (à implémenter selon vos besoins)
            pass
    
    # Tri par date décroissante
    pointages = pointages.order_by('-date', 'employee__full_name')
    
    # Calcul des statistiques
    total_pointages = pointages.count()
    
    # Calculer les statistiques avec des filtres séparés
    present_count = pointages.filter(
        heure_arrivee_minutes__lte=retard_minutes,
        heure_travaillee_secondes__gte=heures_travail_normales_secondes
    ).count()
    retards_count = pointages.filter(heure_arrivee_minutes__gt=retard_minutes).count()
    heures_insuffisantes_count = pointages.filter(heure_travaillee_secondes__lt=heures_travail_normales_secondes).count()
    
    # Pagination
    paginator = Paginator(pointages, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'query': query,
        'date_debut': date_debut,
        'date_fin': date_fin,
        'status_filter': status_filter,
        'heure_arrivee_normale': time(9, 0),
        'heure_retard': time(9, 10),
        'heures_travail_normales': timedelta(hours=8),
        'total_pointages': total_pointages,
        'present_count': present_count,
        'retards_count': retards_count,
        'heures_insuffisantes_count': heures_insuffisantes_count,
        'retard_minutes': retard_minutes,
        'heures_travail_normales_secondes': heures_travail_normales_secondes,
    }
    
    return render(request, 'suivi.html', context)

def employee_pointages(request, employee_id):
    """
    Vue pour afficher tous les pointages d'un employé spécifique
    """
    from datetime import datetime, time, timedelta
    from django.db.models import Q, F, ExpressionWrapper, fields, Count, Case, When, Value, IntegerField
    from django.db.models.functions import ExtractHour, ExtractMinute, Extract
    
    try:
        employee = Employee.objects.get(id=employee_id)
    except Employee.DoesNotExist:
        messages.error(request, 'Employé non trouvé.')
        return redirect('recruitment_suivi')
    
    # Récupération des paramètres de filtrage
    date_debut = request.GET.get('dateDebut', '')
    date_fin = request.GET.get('dateFin', '')
    
    # Base queryset pour cet employé
    pointages = Pointage.objects.filter(employee=employee)
    
    # Filtrage par date
    if date_debut:
        try:
            debut = datetime.strptime(date_debut, '%Y-%m-%d').date()
            pointages = pointages.filter(date__gte=debut)
        except ValueError:
            pass
    
    if date_fin:
        try:
            fin = datetime.strptime(date_fin, '%Y-%m-%d').date()
            pointages = pointages.filter(date__lte=fin)
        except ValueError:
            pass
    
    # Règles métier
    heure_arrivee_normale_minutes = 9 * 60  # 9h00 = 540 minutes
    retard_minutes = (9 * 60) + 10  # 9h10 = 550 minutes
    heures_travail_normales_secondes = 8 * 3600  # 8h = 28800 secondes
    
    # Annoter les pointages avec les statuts
    pointages = pointages.annotate(
        heure_arrivee_minutes=ExpressionWrapper(
            ExtractHour('heure_arrivee') * 60 + ExtractMinute('heure_arrivee'),
            output_field=fields.IntegerField()
        ),
        heure_travaillee_secondes=ExpressionWrapper(
            Extract('heure_travaillee', 'epoch'),
            output_field=fields.IntegerField()
        )
    )
    
    # Tri par date décroissante
    pointages = pointages.order_by('-date')
    
    # Calcul des statistiques pour cet employé
    total_pointages = pointages.count()
    
    present_count = pointages.filter(
        heure_arrivee_minutes__lte=retard_minutes,
        heure_travaillee_secondes__gte=heures_travail_normales_secondes
    ).count()
    retards_count = pointages.filter(heure_arrivee_minutes__gt=retard_minutes).count()
    heures_insuffisantes_count = pointages.filter(heure_travaillee_secondes__lt=heures_travail_normales_secondes).count()
    
    # Pagination
    paginator = Paginator(pointages, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'employee': employee,
        'page_obj': page_obj,
        'date_debut': date_debut,
        'date_fin': date_fin,
        'heure_arrivee_normale': time(9, 0),
        'heure_retard': time(9, 10),
        'heures_travail_normales': timedelta(hours=8),
        'total_pointages': total_pointages,
        'present_count': present_count,
        'retards_count': retards_count,
        'heures_insuffisantes_count': heures_insuffisantes_count,
        'retard_minutes': retard_minutes,
        'heures_travail_normales_secondes': heures_travail_normales_secondes,
    }
    
    return render(request, 'employee_pointages.html', context)

def recruitment_evaluation(request):
    """Vue pour la page d'évaluation des employés"""
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'create_evaluation':
            try:
                # Récupération des données du formulaire
                employee_id = request.POST.get('employee')
                evaluator_id = request.POST.get('evaluator')
                work_quality = int(request.POST.get('work_quality', 3))
                deadline_respect = int(request.POST.get('deadline_respect', 3))
                team_spirit = int(request.POST.get('team_spirit', 3))
                autonomy = int(request.POST.get('autonomy', 3))
                initiative = int(request.POST.get('initiative', 3))
                comments = request.POST.get('comments', '')
                improvement_plan = request.POST.get('improvement_plan', '')
                # Validation des champs requis
                if not employee_id:
                    messages.error(request, 'Veuillez sélectionner un employé.')
                    return redirect('recruitment_evaluation')
                if not evaluator_id:
                    messages.error(request, 'Veuillez sélectionner un évaluateur.')
                    return redirect('recruitment_evaluation')
                overall_score = (work_quality + deadline_respect + team_spirit + autonomy + initiative) / 5
                # Statut selon la note
                if overall_score > 3.75:  # > 7.5/10
                    status = 'validated'
                else:
                    status = 'draft'
                evaluation = Evaluation.objects.create(
                    employee_id=employee_id,
                    evaluator_id=evaluator_id,
                    work_quality=work_quality,
                    deadline_respect=deadline_respect,
                    team_spirit=team_spirit,
                    autonomy=autonomy,
                    initiative=initiative,
                    comments=comments,
                    improvement_plan=improvement_plan,
                    overall_score=overall_score,
                    status=status
                )
                messages.success(request, 'Évaluation créée avec succès.')
                return redirect('recruitment_evaluation')
            except Exception as e:
                messages.error(request, f'Erreur lors de la création de l\'évaluation: {str(e)}')
                return redirect('recruitment_evaluation')
        
        elif action == 'edit_evaluation':
            try:
                evaluation_id = request.POST.get('evaluation_id')
                evaluation = Evaluation.objects.get(id=evaluation_id)
                
                # Mise à jour des champs
                evaluation.work_quality = int(request.POST.get('work_quality', evaluation.work_quality))
                evaluation.deadline_respect = int(request.POST.get('deadline_respect', evaluation.deadline_respect))
                evaluation.team_spirit = int(request.POST.get('team_spirit', evaluation.team_spirit))
                evaluation.autonomy = int(request.POST.get('autonomy', evaluation.autonomy))
                evaluation.initiative = int(request.POST.get('initiative', evaluation.initiative))
                evaluation.comments = request.POST.get('comments', evaluation.comments)
                evaluation.improvement_plan = request.POST.get('improvement_plan', evaluation.improvement_plan)
                
                # Recalculer la note globale
                evaluation.overall_score = (
                    evaluation.work_quality + 
                    evaluation.deadline_respect + 
                    evaluation.team_spirit + 
                    evaluation.autonomy + 
                    evaluation.initiative
                ) / 5
                
                # Mettre à jour le statut selon la nouvelle note
                if evaluation.overall_score > 3.75:  # > 7.5/10
                    evaluation.status = 'validated'
                else:
                    evaluation.status = 'draft'
                
                evaluation.save()
                messages.success(request, 'Évaluation modifiée avec succès.')
                
            except Evaluation.DoesNotExist:
                messages.error(request, 'Évaluation non trouvée.')
            except Exception as e:
                messages.error(request, f'Erreur lors de la modification: {str(e)}')
        
        elif action == 'update_status':
            try:
                evaluation_id = request.POST.get('evaluation_id')
                new_status = request.POST.get('new_status')
                
                evaluation = Evaluation.objects.get(id=evaluation_id)
                evaluation.status = new_status
                
                if new_status == 'validated':
                    evaluation.hr_validation = True
                    evaluation.hr_validator = request.user.employee if hasattr(request.user, 'employee') else None
                    evaluation.hr_validation_date = timezone.now()
                
                evaluation.save()
                messages.success(request, 'Statut de l\'évaluation mis à jour.')
                
            except Exception as e:
                messages.error(request, f'Erreur lors de la mise à jour: {str(e)}')
        
        elif action == 'delete_evaluation':
            try:
                evaluation_id = request.POST.get('evaluation_id')
                evaluation = Evaluation.objects.get(id=evaluation_id)
                evaluation.delete()
                messages.success(request, 'Évaluation supprimée avec succès.')
                
            except Evaluation.DoesNotExist:
                messages.error(request, 'Évaluation non trouvée.')
            except Exception as e:
                messages.error(request, f'Erreur lors de la suppression: {str(e)}')
        
        elif action == 'notify_employee':
            try:
                evaluation_id = request.POST.get('evaluation_id')
                evaluation = Evaluation.objects.get(id=evaluation_id)
                evaluation.employee_notified = True
                evaluation.notification_date = timezone.now()
                evaluation.save()
                messages.success(request, 'Salarié notifié avec succès.')
                
            except Exception as e:
                messages.error(request, f'Erreur lors de la notification: {str(e)}')
    
    # Récupération des données pour l'affichage
    departments = Department.objects.all().order_by('name')
    employees = Employee.objects.filter(status='A').order_by('full_name')
    evaluations = Evaluation.objects.all().order_by('-evaluation_date')
    
    context = {
        'departments': departments,
        'employees': employees,
        'evaluations': evaluations,
        'evaluation_status_choices': Evaluation.STATUS_CHOICES,
    }
    
    return render(request, 'evaluation.html', context)

def download_evaluation_pdf(request, evaluation_id):
    try:
        evaluation = Evaluation.objects.select_related('employee', 'evaluator', 'employee__department').get(id=evaluation_id)
        buffer = io.BytesIO()
        p = canvas.Canvas(buffer)
        y = 800
        p.setFont("Helvetica-Bold", 16)
        p.drawString(100, y, "Fiche d'évaluation")
        y -= 40
        p.setFont("Helvetica", 12)
        p.drawString(100, y, f"Nom complet: {evaluation.employee.full_name}")
        y -= 20
        p.drawString(100, y, f"Département: {evaluation.employee.department.name if evaluation.employee.department else ''}")
        y -= 20
        p.drawString(100, y, f"Évaluateur: {evaluation.evaluator.full_name if evaluation.evaluator else ''}")
        y -= 30
        p.setFont("Helvetica-Bold", 13)
        p.drawString(100, y, "Grille d'évaluation :")
        y -= 20
        p.setFont("Helvetica", 12)
        p.drawString(100, y, f"Qualité du travail: {evaluation.work_quality}/5")
        y -= 20
        p.drawString(100, y, f"Respect des délais: {evaluation.deadline_respect}/5")
        y -= 20
        p.drawString(100, y, f"Esprit d'équipe: {evaluation.team_spirit}/5")
        y -= 20
        p.drawString(100, y, f"Autonomie: {evaluation.autonomy}/5")
        y -= 20
        p.drawString(100, y, f"Capacité d'initiative: {evaluation.initiative}/5")
        y -= 30
        p.drawString(100, y, f"Note globale: {evaluation.overall_score}/5")
        y -= 30
        p.drawString(100, y, f"Commentaires: {evaluation.comments or ''}")
        y -= 20
        p.drawString(100, y, f"Plan d'amélioration: {evaluation.improvement_plan or ''}")
        p.showPage()
        p.save()
        buffer.seek(0)
        filename = f"evaluation_{evaluation.employee.full_name.replace(' ', '_')}.pdf"
        return FileResponse(buffer, as_attachment=True, filename=filename)
    except Exception as e:
        return HttpResponse(f"Erreur lors de la génération du PDF: {e}", status=500)


@csrf_exempt
def cv_upload_parse(request):
    """Handle CV upload and parsing"""
    if request.method == 'POST':
        try:
            from .forms import SingleCVUploadForm, MultipleCVUploadForm
            from .services.cv_parser import CVParser
            import os
            import tempfile
            
            parser = CVParser()
            candidates_data = []
            
            # Check if it's a single CV or multiple CVs
            if 'cv_file' in request.FILES:
                # Single CV upload
                form = SingleCVUploadForm(request.POST, request.FILES)
                if form.is_valid():
                    cv_file = form.cleaned_data['cv_file']
                    
                    # Save to temporary location
                    with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as tmp_file:
                        for chunk in cv_file.chunks():
                            tmp_file.write(chunk)
                        tmp_path = tmp_file.name
                    
                    try:
                        # Parse the CV
                        candidate_data = parser.parse_single_cv(tmp_path)
                        candidates_data.append(candidate_data)
                    finally:
                        # Clean up temporary file
                        if os.path.exists(tmp_path):
                            os.unlink(tmp_path)
                            
            elif 'cv_files' in request.FILES:
                # Multiple CVs upload - completely bypass Django form validation
                try:
                    cv_files = request.FILES.getlist('cv_files')
                    
                    if not cv_files:
                        return JsonResponse({'error': 'Aucun fichier reçu'}, status=400)
                    
                    # Process each PDF file
                    candidates_data = []
                    total_size = 0
                    
                    for cv_file in cv_files:
                        # Validate file type and size
                        if not cv_file.name.lower().endswith('.pdf'):
                            continue  # Skip non-PDF files
                        
                        if cv_file.size > 10 * 1024 * 1024:  # 10MB per file limit
                            continue  # Skip oversized files
                        
                        total_size += cv_file.size
                        if total_size > 50 * 1024 * 1024:  # 50MB total limit
                            break  # Stop if total size exceeds limit
                        
                        # Save to temporary location
                        with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as tmp_file:
                            for chunk in cv_file.chunks():
                                tmp_file.write(chunk)
                            tmp_path = tmp_file.name
                        
                        try:
                            # Parse the CV
                            candidate_data = parser.parse_single_cv(tmp_path)
                            candidates_data.append(candidate_data)
                        finally:
                            # Clean up temporary file
                            if os.path.exists(tmp_path):
                                os.unlink(tmp_path)
                except Exception as e:
                    return JsonResponse({'error': f'Erreur lors du traitement des fichiers multiples: {str(e)}'}, status=500)
            else:
                return JsonResponse({'error': 'Aucun fichier reçu'}, status=400)
            
            # Form validation is handled individually for single uploads only
            
            # Validate extracted data
            valid_candidates = []
            for candidate in candidates_data:
                is_valid, errors = parser.validate_candidate_data(candidate)
                if is_valid:
                    valid_candidates.append(candidate)
                else:
                    # Add errors to candidate data for display
                    candidate['errors'] = errors
                    valid_candidates.append(candidate)
            
            return JsonResponse({
                'success': True,
                'candidates': valid_candidates,
                'total_parsed': len(candidates_data),
                'valid_count': len([c for c in valid_candidates if not c.get('errors')])
            })
            
        except Exception as e:
            return JsonResponse({'error': f'Erreur lors du traitement: {str(e)}'}, status=500)
    
    return JsonResponse({'error': 'Méthode non autorisée'}, status=405)

@csrf_exempt
def save_parsed_candidates(request):
    """Save parsed candidates to database"""
    if request.method == 'POST':
        try:
            import json
            from .models import Candidate
            from django.core.files.base import ContentFile
            import base64
            import uuid
            
            data = json.loads(request.body)
            candidates_data = data.get('candidates', [])
            
            saved_candidates = []
            errors = []
            
            def _normalize_moroccan_phone(raw):
                try:
                    if not raw:
                        return ''
                    digits = re.sub(r'\D', '', str(raw))
                    if not digits:
                        return ''
                    if digits.startswith('00212'):
                        digits = digits[2:]
                    if digits.startswith('212') and len(digits) >= 12:
                        return '+212' + digits[3:12]
                    if digits.startswith('0') and len(digits) >= 10:
                        return '+212' + digits[1:10]
                    if digits.startswith('212') and len(digits) == 12:
                        return '+212' + digits[3:]
                    return ''
                except Exception:
                    return ''

            for candidate_data in candidates_data:
                try:
                    # Check if candidate already exists (by email)
                    if candidate_data.get('email'):
                        existing_candidate = Candidate.objects.filter(email=candidate_data['email']).first()
                        if existing_candidate:
                            errors.append(f"Candidat avec l'email {candidate_data['email']} existe déjà")
                            continue
                    
                    # Normalize phone
                    phone_norm = _normalize_moroccan_phone(candidate_data.get('phone', ''))
                    
                    # Create new candidate
                    candidate = Candidate.objects.create(
                        last_name=candidate_data.get('last_name', ''),
                        first_name=candidate_data.get('first_name', ''),
                        email=candidate_data.get('email', ''),
                        phone=phone_norm,
                        birth_date=candidate_data.get('birth_date'),
                        address=candidate_data.get('address', ''),
                        gender=candidate_data.get('gender', ''),
                        linkedin_profile=candidate_data.get('linkedin_profile', '')
                    )

                    # Save extracted photo if present (data URL base64)
                    photo_data_url = candidate_data.get('photo')
                    if isinstance(photo_data_url, str) and photo_data_url.startswith('data:image') and ';base64,' in photo_data_url:
                        try:
                            header, b64data = photo_data_url.split(';base64,', 1)
                            ext = 'jpg'
                            if 'image/png' in header:
                                ext = 'png'
                            elif 'image/webp' in header:
                                ext = 'webp'
                            # decode and save
                            binary = base64.b64decode(b64data)
                            filename = f"cv_extracted_{uuid.uuid4().hex}.{ext}"
                            candidate.photo.save(filename, ContentFile(binary), save=True)
                        except Exception:
                            # Ignore photo save errors but continue
                            pass
                    
                    saved_candidates.append({
                        'id': candidate.id,
                        'name': f"{candidate.first_name} {candidate.last_name}",
                        'email': candidate.email
                    })
                    
                except Exception as e:
                    errors.append(f"Erreur lors de la sauvegarde de {candidate_data.get('first_name', '')} {candidate_data.get('last_name', '')}: {str(e)}")
            
            return JsonResponse({
                'success': True,
                'saved_count': len(saved_candidates),
                'saved_candidates': saved_candidates,
                'errors': errors
            })
            
        except Exception as e:
            return JsonResponse({'error': f'Erreur lors de la sauvegarde: {str(e)}'}, status=500)
    
    return JsonResponse({'error': 'Méthode non autorisée'}, status=405)
